from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from io import BytesIO
from pathlib import Path
from urllib.parse import quote
import json
import re

from routers.students import filter_students

router = APIRouter()

BASE_DIR = Path("data/exams")
SUBJECTS_PATH = Path("data/subjects.json")


@router.get("/exams/export")
def export_exam_scores(course: str, grade: int, class_name: str, year: str, exam_key: str):

    # 1. Seleciona template dinâmico por grade
    template_file = f"exam_template_{grade}.xlsm"
    TEMPLATE_PATH = Path("templates") / template_file

    if not TEMPLATE_PATH.exists():
        raise HTTPException(404, f"Template Excel para {grade}年 não encontrado.")

    wb = load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb["素点入力（テスト）"]

    # ------------------------------------------------------------------
    # 1.1 ABA 印刷 → B1 DINÂMICO
    # ------------------------------------------------------------------
    ws_print = wb["印刷"]

    EXAM_TITLE_MAP = {
        "前期中間": "前期中間考査",
        "前期期末": "前期期末考査",
        "後期中間": "後期中間考査",
        "後期期末": "後期期末考査",
        "単位認定考査": "単位認定考査"
    }

    exam_title = EXAM_TITLE_MAP.get(exam_key, exam_key)

    # você já tem o school year → year
    school_year = year

    ws_print["B1"] = f"{school_year}年度 {exam_title}"

    # ------------------------------------------------------------------
    # 2. Carrega JSON das provas
    # ------------------------------------------------------------------
    filename = f"{year}-{course}-{grade}-{class_name}.json"
    json_path = BASE_DIR / filename

    if not json_path.exists():
        raise HTTPException(404, "Arquivo de provas não encontrado.")

    with open(json_path, "r", encoding="utf-8") as f:
        exam_data = json.load(f)

    # ------------------------------------------------------------------
    # 3. Carrega alunos
    # ------------------------------------------------------------------
    students = filter_students(
        grade=str(grade),
        course=course,
        gender=None,
        class_name=class_name
    )

    students = [s for s in students if s.get("status") != "休学"]

    if not students:
        raise HTTPException(404, "Nenhum aluno encontrado para esta turma.")

    students = sorted(students, key=lambda s: (s.get("attend_no") or ""))

    # ------------------------------------------------------------------
    # 4. CABEÇALHO: GRADE E CLASS_NAME EM TODAS AS LINHAS (B e C)
    # ------------------------------------------------------------------

    if ws.merged_cells:
        for merged in list(ws.merged_cells):
            ws.unmerge_cells(str(merged))

    grade_num = int(grade)

    match = re.search(r"\d+", class_name)
    class_num = int(match.group()) if match else class_name

    row = 3
    for _st in students:
        ws[f"B{row}"] = grade_num
        ws[f"C{row}"] = class_num
        row += 1

    # ------------------------------------------------------------------
    # 5. PREENCHE ALUNOS (D e E)
    # ------------------------------------------------------------------
    row = 3
    for st in students:
        ws[f"D{row}"] = st.get("attend_no") or ""
        ws[f"E{row}"] = st.get("name") or ""
        row += 1

    # ------------------------------------------------------------------
    # 6. CARREGA SUBJECTS E MONTA LISTA DE MATÉRIAS
    # ------------------------------------------------------------------
    with open(SUBJECTS_PATH, "r", encoding="utf-8") as f:
        subjects_list = json.load(f)

    valid_subjects = [
        s for s in subjects_list
        if s.get("course") == course
        and str(s.get("grade")) == str(grade)
        and s.get("type") != "optional"
        and int(s.get("exam_frequency", 0)) > 0
    ]

    SUBJECT_ORDER = ["国語", "公民", "地理歴史", "数学", "理科",
                     "英語", "保健体育", "芸術", "家庭", "情報", "総合探究"]

    def sort_key(s):
        main = SUBJECT_ORDER.index(s["name"]) if s["name"] in SUBJECT_ORDER else 999
        return (main, s["subject_group"])

    valid_subjects.sort(key=sort_key)

    # ------------------------------------------------------------------
    # 7. MAPA DE EXAM_KEY (LABEL JP → CHAVE INTERNA DO JSON)
    # ------------------------------------------------------------------
    EXAM_KEY_MAP = {
        "前期中間": "zenki_chukan",
        "前期期末": "zenki_kimatsu",
        "後期中間": "koki_chukan",
        "後期期末": "koki_kimatsu",
        "単位認定考査": "single_exam"
    }
    real_key = EXAM_KEY_MAP.get(exam_key, exam_key)

    # ------------------------------------------------------------------
    # 8. PREENCHE TODAS AS MATÉRIAS (HEADER = subject_group, NOTAS = JSON)
    # ------------------------------------------------------------------
    col = 7  # G

    for subj in valid_subjects:
        subject_id = subj["id"]
        subject_group = subj["subject_group"]

        ws.cell(row=2, column=col).value = subject_group

        row = 3
        subject_exams = exam_data.get(subject_id, {})

        for st in students:
            sid = st["id"]
            score = subject_exams.get(real_key, {}).get(sid, "")
            ws.cell(row=row, column=col).value = score
            row += 1

        col += 1

    # ------------------------------------------------------------------
    # 9. Salva e retorna
    # ------------------------------------------------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    out_name = f"{course}-{grade}-{class_name}-{exam_key}.xlsm"
    safe = quote(out_name)

    return StreamingResponse(
        output,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe}"}
    )
